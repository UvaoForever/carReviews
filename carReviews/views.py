import csv
import openpyxl

from django.http import HttpResponse
from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.authentication import TokenAuthentication

from carReviews import models, serializers
from .permissions import IsAdminOrReadOnly, IsUserOrAdmin

def insert_default_countries():
    counties = ["Япония", "Южная Корея", "Германия", "Франция", "США"]
    for country in counties:
        new_country = models.Country.objects.update_or_create(name=country)

def insert_default_producers():
    producer = models.Producer.objects.update_or_create(name="BMW", country_id=3)
    producer = models.Producer.objects.update_or_create(name="Daimler", country_id=3)
    producer = models.Producer.objects.update_or_create(name="Suzuki", country_id=1)
    producer = models.Producer.objects.update_or_create(name="Groupe PSA", country_id=4)
    producer = models.Producer.objects.update_or_create(name="FCA", country_id=5)
    producer = models.Producer.objects.update_or_create(name="Honda Motor", country_id=1)
    producer = models.Producer.objects.update_or_create(name="Ford Motor", country_id=5)
    producer = models.Producer.objects.update_or_create(name="VAG", country_id=3)

def insert_default_autos():
    auto = models.Auto.objects.update_or_create(name="BMW", producer_id=1, start_year=1978, end_year=1989)
    auto = models.Auto.objects.update_or_create(name="Rolls-Royce", producer_id=1, start_year=2013)
    auto = models.Auto.objects.update_or_create(name="Mercedes-Benz", producer_id=2, start_year=2022)
    auto = models.Auto.objects.update_or_create(name="Suzuki", producer_id=3, start_year=2016)
    auto = models.Auto.objects.update_or_create(name="Peugeot", producer_id=4, start_year=1999, end_year=2017)
    auto = models.Auto.objects.update_or_create(name="Citroen", producer_id=4, start_year=1974, end_year=1991)
    auto = models.Auto.objects.update_or_create(name="Ferrari", producer_id=5, start_year=2012, end_year=2017)
    auto = models.Auto.objects.update_or_create(name="Honda", producer_id=6, start_year=2014, end_year=2016)
    auto = models.Auto.objects.update_or_create(name="Ford", producer_id=7, start_year=2022)
    auto = models.Auto.objects.update_or_create(name="Audi", producer_id=8, start_year=2010, end_year=2018)

def insert_default_comments():
    comment = models.Comment.objects.update_or_create(email="petr1980@mail.ru", auto_id=1, date_comment="2022-12-04", comment="Очень доволен машиной")
    comment = models.Comment.objects.update_or_create(email="sendbox@google.com", auto_id=2, date_comment="2024-10-23", comment="Машина хорошая, рекомендую")
    comment = models.Comment.objects.update_or_create(email="evgen@google.com", auto_id=3, date_comment="2023-11-02", comment="Категорически не рекомендую")
    comment = models.Comment.objects.update_or_create(email="pavel2302@mail.ru", auto_id=4, date_comment="2020-10-23", comment="Жаль, что больше не выпускают. Хорошая машина")
    comment = models.Comment.objects.update_or_create(email="aria@mail.ru", auto_id=5, date_comment="2019-05-15", comment="Пока сложность сказать, до конца не опробовал")
    comment = models.Comment.objects.update_or_create(email="arsen20@google.com", auto_id=6, date_comment="2021-07-21", comment="Купил в подарок сыну. Всё нравится")
    comment = models.Comment.objects.update_or_create(email="konami@mail.ru", auto_id=7, date_comment="2018-12-23", comment="Только что приобрёл. Пока никаких недостатков не заметил")
    comment = models.Comment.objects.update_or_create(email="akira1998@mail.ru", auto_id=8, date_comment="2022-10-10", comment="Меня всё устраивает")

def insert_default_data():
    insert_default_countries()
    insert_default_producers()
    insert_default_autos()
    insert_default_comments()

class CountryViewSet(viewsets.ModelViewSet):
    # queryset = models.Country.objects.all()
    serializer_class = serializers.CountrySerializer
    permission_classes = (IsAdminOrReadOnly,)

    def get_queryset(self):
        pk = self.kwargs.get("pk")

        if not pk:
            return models.Country.objects.all()

        return models.Country.objects.filter(pk=pk)

class ProducerViewSet(viewsets.ModelViewSet):
    # queryset = models.Country.objects.all()
    serializer_class = serializers.ProducerSerializer
    permission_classes = (IsAdminOrReadOnly,)

    def get_queryset(self):
        pk = self.kwargs.get("pk")

        if not pk:
            return models.Producer.objects.all()

        return models.Producer.objects.filter(pk=pk)

class AutoViewSet(viewsets.ModelViewSet):
    # queryset = models.Country.objects.all()
    serializer_class = serializers.AutoSerializer
    permission_classes = (IsAdminOrReadOnly, )
    authentication_classes = (TokenAuthentication, )

    def get_queryset(self):
        pk = self.kwargs.get("pk")

        if not pk:
            return models.Auto.objects.all()

        return models.Auto.objects.filter(pk=pk)

class CommentViewSet(viewsets.ModelViewSet):
    serializer_class = serializers.CommentSerializer
    permission_classes = (IsUserOrAdmin, )
    authentication_classes = (TokenAuthentication, )

    def get_queryset(self):
        pk = self.kwargs.get("pk")

        if not pk:
            return models.Comment.objects.all()

        return models.Comment.objects.filter(pk=pk)

# Экспорт стран
def parsing_countries(request):
    countries = models.Country.objects.all()
    file_format = request.GET.get('format', 'csv')  # По умолчанию экспортация в CSV

    if file_format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=windows-1251')
        response['Content-Disposition'] = 'attachment; filename="countries.xlsx"'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Countries'

        # Заголовки столбцов
        sheet.append(['id', 'country'])

        for country in countries:
            sheet.append([country.id, country.name])

        workbook.save(response)
        return response

    response = HttpResponse(content_type='text/csv; charset=windows-1251')
    response['Content-Disposition'] = 'attachment; filename="countries.csv"'
    writer = csv.writer(response)
    # Заголовки столбцов
    writer.writerow(['id', 'country'])
    for country in countries:
        writer.writerow([country.id, country.name])

    return response

# Экспорт производителей
def parsing_producers(request):
    producers = models.Producer.objects.all()

    file_format = request.GET.get('format', 'csv')  # По умолчанию экспортация в CSV

    if file_format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=windows-1251')
        response['Content-Disposition'] = 'attachment; filename="producers.xlsx"'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Producers'

        # Заголовки столбцов
        sheet.append(['id', 'name', "country"])

        for producer in producers:
            sheet.append([producer.id, producer.name, producer.country.name])

        workbook.save(response)
        return response

    response = HttpResponse(content_type='text/csv; charset=windows-1251')
    response['Content-Disposition'] = 'attachment; filename="producers.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'name', "country"])
    for producer in producers:
        writer.writerow([producer.id, producer.name, producer.country.name])

    return response

# Экспорт автомобилей
def parsing_autos(request):
    autos = models.Auto.objects.all()

    file_format = request.GET.get('format', 'csv')  # По умолчанию экспортация в CSV

    if file_format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=windows-1251')
        response['Content-Disposition'] = 'attachment; filename="autos.xlsx"'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Autos'

        # Заголовки столбцов
        sheet.append(['id', 'name', "producer", "start_year", "end_year"])

        for auto in autos:
            sheet.append([auto.id, auto.name, auto.producer.name, auto.start_year, auto.end_year])

        workbook.save(response)
        return response

    response = HttpResponse(content_type='text/csv; charset=windows-1251')
    response['Content-Disposition'] = 'attachment; filename="autos.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'name', "producer", "start_year", "end_year"])
    for auto in autos:
        writer.writerow([auto.id, auto.name, auto.producer.name, auto.start_year, auto.end_year])

    return response

# Экспорт комментариев
def parsing_comments(request):
    comments = models.Comment.objects.all()

    file_format = request.GET.get('format', 'csv')  # По умолчанию экспортация в CSV

    if file_format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=windows-1251')
        response['Content-Disposition'] = 'attachment; filename="comments.xlsx"'

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Comments'

        # Заголовки столбцов
        sheet.append(['id', 'email', "auto", "date_comment", "comment"])

        for comment in comments:
            sheet.append([comment.id, comment.email, comment.auto.name, comment.date_comment, comment.comment])

        workbook.save(response)
        return response

    response = HttpResponse(content_type='text/csv; charset=windows-1251')
    response['Content-Disposition'] = 'attachment; filename="comments.csv"'
    writer = csv.writer(response)
    writer.writerow(['id', 'email', "auto", "date_comment", "comment"])
    for comment in comments:
        writer.writerow([comment.id, comment.email, comment.auto.name, comment.date_comment, comment.comment])

    return response

def index(request):
    insert_default_data()
    countries = models.Country.objects.all()
    producers = models.Producer.objects.all()
    cars = models.Auto.objects.all()
    comments = models.Comment.objects.all()
    return render(request, "index.html", {"countries" : countries, "producers" : producers, "cars" : cars, "comments" : comments})